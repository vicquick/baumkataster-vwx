"""VectorWorks Import TXT/XLSX generator."""

import io

import geopandas as gpd
from pyproj import CRS, Transformer

# UTF-8 BOM — helps VectorWorks on Windows detect encoding correctly
_UTF8_BOM = "\ufeff"

LS320_WKT = (
    'PROJCS["ETRS89 / Gauss-Kruger CM 9E (LS320)",'
    'GEOGCS["ETRS89",'
    'DATUM["European_Terrestrial_Reference_System_1989",'
    'SPHEROID["GRS 1980",6378137,298.257222101]],'
    'PRIMEM["Greenwich",0],'
    'UNIT["degree",0.0174532925199433]],'
    'PROJECTION["Transverse_Mercator"],'
    'PARAMETER["latitude_of_origin",0],'
    'PARAMETER["central_meridian",9],'
    'PARAMETER["scale_factor",1],'
    'PARAMETER["false_easting",3500000],'
    'PARAMETER["false_northing",0],'
    'UNIT["metre",1],'
    'AXIS["Easting",EAST],'
    'AXIS["Northing",NORTH]]'
)
LS320_CRS = CRS.from_wkt(LS320_WKT)

CRS_OPTIONS = {
    "EPSG:25832": "ETRS89 / UTM zone 32N",
    "EPSG:31467": "DHDN / 3-degree Gauss-Kruger zone 3 (3xxxxxx)",
    "EPSG:4647": "ETRS89 / UTM zone 32N (zE-N, 32xxxxxx)",
    "LS320": "Hamburg LS320 — ETRS89 / GK CM 9E (FE=3500000)",
}


def resolve_crs(key: str):
    """Return a pyproj-compatible CRS for a given key."""
    if key == "LS320":
        return LS320_CRS
    return key


def estimate_baumhoehe(kronendurchmesser):
    """Estimate tree height from crown diameter using urban allometry.

    Based on typical urban broadleaf relationships:
      Höhe ≈ KD × 1.4 + 1.5
    E.g. KD=5m → ~8.5m, KD=7m → ~11.3m, KD=10m → ~15.5m
    """
    try:
        kd = float(kronendurchmesser)
    except (TypeError, ValueError):
        return ""
    if kd <= 0:
        return ""
    return round(kd * 1.4 + 1.5)


def estimate_stammumfang(kronendurchmesser):
    """Estimate trunk circumference (cm) from crown diameter (m).

    Based on typical urban broadleaf relationships:
      StU ≈ KD × 18 + 20
    E.g. KD=5m → ~110cm, KD=7m → ~146cm, KD=10m → ~200cm
    """
    try:
        kd = float(kronendurchmesser)
    except (TypeError, ValueError):
        return ""
    if kd <= 0:
        return ""
    return round(kd * 18 + 20)


def estimate_ansatzhoehe(hoehe, kronendurchmesser, method="ratio", ratio=0.25):
    """Estimate Ansatzhöhe (crown base height) from tree height and crown diameter.

    Methods:
    - "ratio": Ansatzhöhe = Höhe × ratio  (default 0.25, typical range 0.15-0.40)
    - "kd":    Ansatzhöhe = Höhe - Kronendurchmesser (crude, often overestimates)
    """
    try:
        h = float(hoehe)
    except (TypeError, ValueError):
        return ""

    if method == "kd":
        try:
            kd = float(kronendurchmesser)
            result = max(1.0, h - kd)
        except (TypeError, ValueError):
            result = max(1.0, h * ratio)
    else:
        result = max(1.0, h * ratio)

    return round(result)


def _fmt(f):
    """Format a field value for TXT output."""
    if f is None:
        return ""
    s = str(f)
    return "" if s in ("nan", "None") else s


# VW parameter names (German) — exact match to VW Import Tree Survey dropdown.
# Every parameter is included so VW can auto-map by column header name.
# Empty columns are fine — VW maps them to <Kein>.
VW_COLUMNS = [
    # Coordinates
    "x-Koordinate",
    "y-Koordinate",
    # Identity
    "Baum-ID",
    "Botanischer Name",
    "Deutscher Name",
    "Familie",
    "Herkunft",
    "Invasive Art",
    # Dimensions
    "Höhe",
    "Kronendurchmesser",
    "Krone Nord (N)",
    "Krone Nordost (NO)",
    "Krone Ost (O)",
    "Krone Südost (SO)",
    "Krone Süd (S)",
    "Krone Südwest (SW)",
    "Krone West (W)",
    "Krone Nordwest (NW)",
    "Lichte Höhe",
    "Astansatzhöhe",
    "Ausrichtung Hauptast",
    "Stammumfang",
    # Status
    "Maßnahme",
    "Veteranenbaum",
    "Alter Baum",
    "Pflanzjahr",
    "Letzte Überprüfung am",
    "Standort",
    "Bemerkungen",
    "Form",
    "Struktur",
    "Vitalität",
    "Stammfuß-ø",
    # Custom fields
    "Feld 1",
    "Feld 2",
    "Feld 3",
    "Feld 4",
    "Feld 5",
    "Feld 6",
    "Zusatz 07",
    "Zusatz 08",
    "Zusatz 09",
    "Zusatz 10",
]


# --- Language-dependent value mappings for VW enum fields ---
# VW does ENGLISH string matching internally, even on German installs.
# The dropdown shows German labels but import parser matches English values.

# Maßnahme / Action — map our terms to VW internal values
# VW 2026 docs confirm EXACT values: Retain, Transplant, Remove, Custom
_ACTION_TO_VW_EN = {
    # German → English (exact VW values)
    "Sichern": "Retain",
    "Sichern und Baumpflegemaßnahmen": "Retain",
    "Entfernen": "Remove",
    "Entnehmen - innerhalb 4 Wochen": "Remove",
    "Entnehmen - im nächsten Pflegezyklus": "Remove",
    "Umpflanzen - Am Standort": "Transplant",
    "Umpflanzen - Neuer Standort": "Transplant",
    # Legacy / passthrough
    "Roden": "Remove",
    "Umpflanzen": "Transplant",
    "Retain": "Retain",
    "Remove": "Remove",
    "Transplant": "Transplant",
}

_ACTION_TO_VW_DE = {
    # Legacy / English → German
    "Retain": "Sichern",
    "Remove": "Entfernen",
    "Transplant": "Umpflanzen - Am Standort",
    "Roden": "Entfernen",
    "Umpflanzen": "Umpflanzen - Am Standort",
}

# Vitalität / Condition — VW uses Excellent/Good/Average/Poor internally
_VITALITAET_TO_VW_EN = {
    "0": "Excellent",
    "1": "Good",
    "2": "Average",
    "3": "Poor",
}

_VITALITAET_TO_VW_DE = {
    "0": "0 Krone harmonisch geschlossen, fast kein Totholz in der Krone",
    "1": "1 Kronenmantel an wenigen Stellen zerklüftet, wenig Totholz im Dünnast- und Starkastbereich",
    "2": "2 Vermehrt Totholz, Bildung einer Sekundärkrone",
    "3": "3 Absterben von Ästen, sehr viel Totholz in der Krone",
}

# Column header alternatives — English VW header names for auto-mapping
VW_COLUMNS_EN = [
    "x-Koordinate",
    "y-Koordinate",
    "Tree ID",
    "Botanical Name",
    "Common Name",
    "Family",
    "Origin",
    "Invasive Species",
    "Height",
    "Crown Spread",
    "Crown North (N)",
    "Crown North East (NE)",
    "Crown East (E)",
    "Crown South East (SE)",
    "Crown South (S)",
    "Crown South West (SW)",
    "Crown West (W)",
    "Crown North West (NW)",
    "Clear Height",
    "Crown Base Height",
    "Orientation Main Branch",
    "Circumference",
    "Action",
    "Heritage Tree",
    "Old Tree",
    "Year Planted",
    "Last Inspection Date",
    "Location",
    "Notes",
    "Form",
    "Structure",
    "Condition",
    "Root Collar Diameter",
    "Field 1",
    "Field 2",
    "Field 3",
    "Field 4",
    "Field 5",
    "Field 6",
    "Extra 07",
    "Extra 08",
    "Extra 09",
    "Extra 10",
]


def _map_action(val, lang="en"):
    """Map action value to VW Action/Maßnahme value."""
    s = _fmt(val)
    if not s:
        return ""
    table = _ACTION_TO_VW_EN if lang == "en" else _ACTION_TO_VW_DE
    return table.get(s, s)


def _map_vitalitaet(val, lang="en"):
    """Map vitalitaet value to VW Condition/Vitalitätsstufe value.

    Handles single values (0, 1, 2, 3) and ranges (0-1, 1-2).
    For ranges, uses the worse (higher) value.
    """
    s = _fmt(val).strip().replace("–", "-")
    if not s:
        return ""
    table = _VITALITAET_TO_VW_EN if lang == "en" else _VITALITAET_TO_VW_DE
    # If it's a range like "1-2", take the worse (higher) value
    if "-" in s:
        parts = s.split("-")
        try:
            worst = str(max(int(p.strip()) for p in parts))
            return table.get(worst, s)
        except ValueError:
            pass
    return table.get(s, s)


def _build_vw_row(row, x, y, ansatzhoehe="", standort="", lang="en"):
    """Build a row list matching VW_COLUMNS from a GeoDataFrame row."""
    deutsch = row.get("art_deutsch", "") or row.get("gattung_deutsch", "")
    latein = row.get("art_latein", "") or row.get("gattung_latein", "")

    return [
        # Coordinates
        f"{x:.3f}",
        f"{y:.3f}",
        # Identity
        row.get("baum_id", ""),
        latein,
        deutsch,
        row.get("familie", ""),
        row.get("herkunft", ""),
        row.get("invasive_art", ""),
        # Dimensions
        row.get("baumhoehe", ""),
        row.get("kronendurchmesser", ""),
        row.get("krone_n", ""),
        row.get("krone_no", ""),
        row.get("krone_o", ""),
        row.get("krone_so", ""),
        row.get("krone_s", ""),
        row.get("krone_sw", ""),
        row.get("krone_w", ""),
        row.get("krone_nw", ""),
        row.get("lichte_hoehe", ""),
        ansatzhoehe,
        row.get("ausrichtung_hauptast", ""),
        row.get("stammumfang", ""),
        # Status — mapped to VW internal values
        _map_action(row.get("action", ""), lang=lang),
        row.get("veteranenbaum", ""),
        row.get("alter_baum", ""),
        row.get("pflanzjahr", ""),
        row.get("letzte_ueberpruefung", ""),
        standort,
        row.get("bemerkungen", ""),
        row.get("kronenform", ""),
        row.get("struktur", ""),
        _map_vitalitaet(row.get("vitalitaet", ""), lang=lang),
        row.get("stammfuss_durchmesser", ""),
        # Custom fields — PDF extras that VW has no native field for
        row.get("erhaltung", ""),           # Feld 1: Erhaltungswürdigkeit
        row.get("verkehrssicherheit", ""),   # Feld 2: Verkehrssicherheit
        row.get("schutzstatus", ""),         # Feld 3
        row.get("anzahl_staemme", ""),       # Feld 4
        row.get("ersatzpflanzungen", ""),    # Feld 5
        row.get("feld6", ""),
        row.get("zusatz07", ""),
        row.get("zusatz08", ""),
        row.get("zusatz09", ""),
        row.get("zusatz10", ""),
    ]


def _export_vw_txt(all_rows: list[list], lang="en") -> str:
    """Given a list of row-lists (each matching VW_COLUMNS), drop empty columns and build TXT.

    Prepends UTF-8 BOM so VectorWorks on Windows detects encoding correctly.
    Uses English or German column headers depending on lang.
    """
    # Headers always German (VW auto-maps by German header name)
    # Values use lang for enum fields (Action, Condition)
    n_cols = len(VW_COLUMNS)
    has_data = [False] * n_cols
    for row in all_rows:
        for i in range(n_cols):
            if _fmt(row[i]):
                has_data[i] = True

    # Always keep x/y coordinates
    has_data[0] = True
    has_data[1] = True

    keep = [i for i in range(n_cols) if has_data[i]]

    header = "\t".join(VW_COLUMNS[i] for i in keep)
    lines = [_UTF8_BOM + header]
    for row in all_rows:
        lines.append("\t".join(_fmt(row[i]) for i in keep))

    return "\n".join(lines)


def _export_vw_xlsx(all_rows: list[list], lang="en") -> bytes:
    """Given a list of row-lists (each matching VW_COLUMNS), drop empty columns and build XLSX."""
    import openpyxl

    # Headers always German (VW auto-maps by German header name)
    # Values use lang for enum fields (Action, Condition)
    n_cols = len(VW_COLUMNS)
    has_data = [False] * n_cols
    for row in all_rows:
        for i in range(n_cols):
            if _fmt(row[i]):
                has_data[i] = True

    has_data[0] = True
    has_data[1] = True

    keep = [i for i in range(n_cols) if has_data[i]]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Baumkataster"

    # Header row — always German for VW auto-mapping
    for col_idx, i in enumerate(keep, 1):
        ws.cell(row=1, column=col_idx, value=VW_COLUMNS[i])

    # Data rows — try to write numbers as numbers for VW
    for row_idx, row in enumerate(all_rows, 2):
        for col_idx, i in enumerate(keep, 1):
            val = _fmt(row[i])
            if val:
                try:
                    val = float(val)
                except ValueError:
                    pass
            else:
                val = ""
            ws.cell(row=row_idx, column=col_idx, value=val)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _build_wfs_rows(trees_gdf, output_crs_key, ansatz_method="none", ansatz_ratio=0.25,
                    lang="en", estimate_from_kd=False):
    """Build VW row data from WFS/REST GeoDataFrame."""
    transformer = Transformer.from_crs("EPSG:4326", resolve_crs(output_crs_key), always_xy=True)
    all_rows = []
    for _, row in trees_gdf.iterrows():
        x, y = transformer.transform(row.geometry.x, row.geometry.y)
        hoehe = row.get("baumhoehe", "")
        kd = row.get("kronendurchmesser", "")
        stu = row.get("stammumfang", "")

        # Estimate missing dimensions from KD if enabled
        if estimate_from_kd and _fmt(kd):
            if not _fmt(hoehe):
                hoehe = estimate_baumhoehe(kd)
                row = dict(row)
                row["baumhoehe"] = hoehe
            if not _fmt(stu):
                stu = estimate_stammumfang(kd)
                row = dict(row)
                row["stammumfang"] = stu

        if ansatz_method == "none":
            ansatzhoehe = row.get("ansatzhoehe", "") if hasattr(row, "get") else row.get("ansatzhoehe", "")
        else:
            ansatzhoehe = estimate_ansatzhoehe(hoehe, kd, method=ansatz_method, ratio=ansatz_ratio)
        strasse = row.get("strasse", "")
        hausnr = row.get("hausnummer", "")
        standort = f"{strasse} {hausnr}".strip() if strasse else ""
        all_rows.append(_build_vw_row(row, x, y, ansatzhoehe=ansatzhoehe, standort=standort, lang=lang))
    return all_rows


def _build_pdf_rows(trees_gdf, output_crs_key, ansatz_method="none", ansatz_ratio=0.25, lang="en"):
    """Build VW row data from PDF-merged GeoDataFrame."""
    transformer = Transformer.from_crs("EPSG:4326", resolve_crs(output_crs_key), always_xy=True)
    all_rows = []
    for _, row in trees_gdf.iterrows():
        x, y = transformer.transform(row.geometry.x, row.geometry.y)
        ansatzhoehe = row.get("ansatzhoehe", "")
        if not _fmt(ansatzhoehe) and ansatz_method != "none":
            hoehe = row.get("baumhoehe", "")
            kd = row.get("kronendurchmesser", "")
            ansatzhoehe = estimate_ansatzhoehe(hoehe, kd, method=ansatz_method, ratio=ansatz_ratio)
        standort = row.get("standort", "")
        all_rows.append(_build_vw_row(row, x, y, ansatzhoehe=ansatzhoehe, standort=standort, lang=lang))
    return all_rows


# --- TXT exports ---

def trees_to_vw_txt(trees_gdf, output_crs_key, ansatz_method="none", ansatz_ratio=0.25,
                    lang="en", estimate_from_kd=False):
    """WFS/REST pipeline → VW import TXT."""
    rows = _build_wfs_rows(trees_gdf, output_crs_key, ansatz_method, ansatz_ratio,
                           lang=lang, estimate_from_kd=estimate_from_kd)
    return _export_vw_txt(rows, lang=lang)


def pdf_trees_to_vw_txt(trees_gdf, output_crs_key, ansatz_method="none", ansatz_ratio=0.25, lang="en"):
    """PDF pipeline → VW import TXT."""
    rows = _build_pdf_rows(trees_gdf, output_crs_key, ansatz_method, ansatz_ratio, lang=lang)
    return _export_vw_txt(rows, lang=lang)


# --- XLSX exports ---

def trees_to_vw_xlsx(trees_gdf, output_crs_key, ansatz_method="none", ansatz_ratio=0.25,
                     lang="en", estimate_from_kd=False):
    """WFS/REST pipeline → VW import XLSX."""
    rows = _build_wfs_rows(trees_gdf, output_crs_key, ansatz_method, ansatz_ratio,
                           lang=lang, estimate_from_kd=estimate_from_kd)
    return _export_vw_xlsx(rows, lang=lang)


def pdf_trees_to_vw_xlsx(trees_gdf, output_crs_key, ansatz_method="none", ansatz_ratio=0.25, lang="en"):
    """PDF pipeline → VW import XLSX."""
    rows = _build_pdf_rows(trees_gdf, output_crs_key, ansatz_method, ansatz_ratio, lang=lang)
    return _export_vw_xlsx(rows, lang=lang)


# --- VW Python script workaround for import bugs ---

def generate_fixup_script(trees_gdf, output_crs_key, lang="en",
                          ansatz_method="none", ansatz_ratio=0.25):
    """Generate a VW Python script to batch-set ALL fields on Existing Tree objects.

    Works around VW import bugs where fields don't populate correctly.
    Matches imported trees by coordinate proximity (since Tree ID may also fail).
    Run this script inside VW's Script Editor after importing the tree survey.
    """
    transformer = Transformer.from_crs("EPSG:4326", resolve_crs(output_crs_key), always_xy=True)

    tree_entries = []
    for _, row in trees_gdf.iterrows():
        x, y = transformer.transform(row.geometry.x, row.geometry.y)
        baum_id = _fmt(row.get("baum_id", ""))
        if not baum_id:
            continue

        deutsch = _fmt(row.get("art_deutsch", "") or row.get("gattung_deutsch", ""))
        latein = _fmt(row.get("art_latein", "") or row.get("gattung_latein", ""))
        action_raw = _fmt(row.get("action", ""))
        action = _map_action(action_raw, lang=lang) if action_raw else ""
        vitalitaet = _map_vitalitaet(row.get("vitalitaet", ""), lang=lang)
        stammumfang = _fmt(row.get("stammumfang", ""))
        kronendurchmesser = _fmt(row.get("kronendurchmesser", ""))
        baumhoehe = _fmt(row.get("baumhoehe", ""))
        kronenform = _fmt(row.get("kronenform", ""))
        struktur = _fmt(row.get("struktur", ""))
        bemerkungen = _fmt(row.get("bemerkungen", ""))
        pflanzjahr = _fmt(row.get("pflanzjahr", ""))
        standort = _fmt(row.get("standort", ""))
        if not standort:
            strasse = _fmt(row.get("strasse", ""))
            hausnr = _fmt(row.get("hausnummer", ""))
            standort = f"{strasse} {hausnr}".strip() if strasse else ""

        ansatzhoehe = _fmt(row.get("ansatzhoehe", ""))
        if not ansatzhoehe and ansatz_method != "none":
            est = estimate_ansatzhoehe(baumhoehe, kronendurchmesser,
                                       method=ansatz_method, ratio=ansatz_ratio)
            ansatzhoehe = str(est) if est != "" else ""

        tree_entries.append({
            "x": round(x, 3), "y": round(y, 3),
            "id": baum_id, "latein": latein, "deutsch": deutsch,
            "action": action, "condition": vitalitaet,
            "stammumfang": stammumfang, "kd": kronendurchmesser,
            "hoehe": baumhoehe, "ansatz": ansatzhoehe,
            "form": kronenform, "struktur": struktur,
            "bemerkungen": bemerkungen, "pflanzjahr": pflanzjahr,
            "standort": standort,
        })

    if not tree_entries:
        return None

    entries_src = "[\n"
    for e in tree_entries:
        entries_src += f"    {e!r},\n"
    entries_src += "]"

    script = f'''# VW Python Script: Batch fix-up ALL fields on Existing Tree objects
# Generated by Baumkataster Tool
# Run in VW: Tools > Scripting > Script Editor > Python
#
# Workaround for VW import bugs (VB-214920, VB-215462, etc.)
# Matches imported trees by coordinate proximity, then sets all fields.
# Auto-discovers actual VW field names — works on any language install.

import vs
import math

TREES = {entries_src}

TOLERANCE = 1.0  # coordinate matching tolerance (document units)

# Exact VW internal field names (discovered from VW 2026 "Existing Tree" record)
FIELD_MAP_EXACT = {{
    "id":           "Tree No",
    "latein":       "Genus Species",
    "deutsch":      "Common Name",
    "action":       "ActionComment",
    "condition":    "Condition",
    "hoehe":        "Height",
    "kd":           "Canopy",
    "stammumfang":  "Calliper",
    "ansatz":       "Hgt First Branch",
    "form":         "Form",
    "bemerkungen":  "Comments",
    "pflanzjahr":   "Year Planted",
    "struktur":     "Structure",
    "standort":     "Location",
}}

count = 0
matched = set()
rec_name = "Existing Tree"

def find_nearest(ox, oy):
    best_i = -1
    best_d = TOLERANCE
    for i, t in enumerate(TREES):
        if i in matched:
            continue
        d = math.sqrt((ox - t["x"])**2 + (oy - t["y"])**2)
        if d < best_d:
            best_d = d
            best_i = i
    return best_i

debug_msg = ""

def fix_tree(h):
    global count, debug_msg

    p = vs.GetSymLoc(h)
    if isinstance(p, tuple) and len(p) >= 2:
        ox, oy = p[0], p[1]
    else:
        return

    idx = find_nearest(ox, oy)
    if idx < 0:
        return
    matched.add(idx)
    t = TREES[idx]

    # Set all non-ID fields
    for key, fname in FIELD_MAP_EXACT.items():
        if key == "id":
            continue
        val = t.get(key, "")
        if val:
            vs.SetRField(h, rec_name, fname, str(val))

    # Tree No: VW computes it from IDPrefix + IDLabel + IDSuffix.
    # Setting "Tree No" directly gets overridden by the PIO on reset.
    # Set the component fields that the PIO uses to build Tree No.
    tree_id = t.get("id", "")
    if tree_id:
        vs.SetRField(h, rec_name, "IDPrefix", "")
        vs.SetRField(h, rec_name, "IDLabel", str(tree_id))
        vs.SetRField(h, rec_name, "IDSuffix", "")
        try:
            vs.SetRField(h, rec_name, "TreeTag", str(tree_id))
        except:
            pass
        vs.SetRField(h, rec_name, "Tree No", str(tree_id))

    # Reset AFTER all fields including ID components are set
    vs.ResetObject(h)

    # Debug: read back Tree No and IDLabel on first tree
    if count == 0:
        rb_treeno = vs.GetRField(h, rec_name, "Tree No")
        rb_label = vs.GetRField(h, rec_name, "IDLabel")
        debug_msg = f"TreeNo='{{rb_treeno}}', IDLabel='{{rb_label}}', wanted='{{tree_id}}'"

    count += 1

# Run on all Existing Tree PIOs
for rn in ["Existing Tree", "Vorhandener Baum"]:
    vs.ForEachObject(fix_tree, f"(T=PLUGINOBJECT) & (R IN ['{{rn}}'])")

vs.AlrtDialog(f"Updated {{count}} of {{len(TREES)}} trees.\\n{{debug_msg}}")
'''
    return script
